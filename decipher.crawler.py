#!/sur/bin python3
#coding='utf-8'
import traceback
import os
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
import time
import re
import logging
import random
from copy import deepcopy
import openpyxl
from openpyxl import Workbook
from selenium.webdriver.remote.remote_connection import LOGGER
from selenium.webdriver.chrome.options import Options
import sys

options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])

# 禁止将日志消息输出到控制台
# LOGGER.setLevel(logging.WARNING)

# 日志
class Logger(object):
    # logging日志格式设置
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s - %(levelname)s: %(message)s')
    @staticmethod
    def info(message: str):
     # info级别的日志，绿色
        logging.info("\033[0;32m" + message + "\033[0m")

    @staticmethod
    def warning(message: str):
     # warning级别的日志，黄色
        logging.warning("\033[0;33m" + message + "\033[0m")

    @staticmethod
    def error(message: str):
     # error级别的日志，红色
        logging.error("\033[0;31m"+"-" * 23 + '\n| ' + message + "\033[0m" + "\n" + "└"+"-" * 55)

    @staticmethod
    def debug(message: str):
     # debug级别的日志，灰色
        logging.debug("\033[0;37m" + message + "\033[0m")

logger = Logger()

def dealing(driver, search_key, num_start, num_end, total_num, get_list):
    logger.info(f"Search DECIPHER {search_key}, get {total_num} records, dealing {num_start} to {num_end} ...")
    # 表头
    result_dicts = []
    headers = []
    table_header = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/table/thead/tr[1]')
    for header_th in table_header.find_elements(by=By.TAG_NAME, value='th'):
        headers.append(header_th.text)
    # print(headers)
    # 表格
    table_body = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/table/tbody')
    for line in table_body.find_elements(by=By.TAG_NAME, value='tr'):
        values = []
        for line_td in line.find_elements(by=By.TAG_NAME, value='td'):
            values.append(line_td.text)
        data_dict = dict(zip(headers, values))
        # 保存一下第一页面的location 便于保存记录
        data_dict['Location_all'] = data_dict['Location']
        data_dict['Location_all'] = re.sub(r'\n', ' ', data_dict['Location_all'])
        #result_dicts.append(data_dict)
        # 跳过已有记录
        get_key = '__'.join([data_dict[key] for key in ['DECIPHER Patient', 'Location_all']])
        if get_key in get_list:
            logger.info(f"{get_key} already in the db, pass it.")
            #print(get_list)
            continue
        else:
            get_list.append(get_key)
        line.find_elements(by=By.TAG_NAME, value='td')[0].click()
        # 点击 patientID 跳转子页 注意需要切换新标签页 获取所需字段
        # 找到对应的变异
        # 获取全部页面句柄
        # 查询库里面是不是有
        all_handles = driver.window_handles
        # 将当前句柄定位到新打开的页面
        driver.switch_to.window(all_handles[-1])
        WebDriverWait(driver, 30, 1, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div/h4'))
        time.sleep(1)
        # 获取表格变异信息 
        headers_1 = []
        table_header_1 = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div[1]/div/table/thead/tr[1]')
        for header_th_1 in table_header_1.find_elements(by=By.TAG_NAME, value='th'):
            headers_1.append(header_th_1.text)
        table_body_1 = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div[1]/div/table/tbody')
        for line_1 in table_body_1.find_elements(by=By.TAG_NAME, value='tr'):
            values_1 = [] 
            # 只匹配chr start end 一样的
            print(line_1.find_elements(by=By.TAG_NAME, value='td')[0].text.strip(), data_dict['Location'])
            if not line_1.find_elements(by=By.TAG_NAME, value='td')[0].text.strip().split('\n')[:5] == data_dict['Location'].strip().split('\n')[:5]:
                continue
            data_dict_new = deepcopy(data_dict)
            for line_td_1 in line_1.find_elements(by=By.TAG_NAME, value='td'):
                values_1.append(line_td_1.text)
            data_dict_1 = dict(zip(headers_1, values_1))
            data_dict_new.update(data_dict_1)
            # 点击弹窗获取坐标
            action_chains = ActionChains(driver)
            # 部分找不到hg19位置
            try:
                action_chains.move_to_element(line_1.find_elements(by=By.TAG_NAME, value='td')[0].find_element(by=By.XPATH, value='.//div/div/a')).click().perform()
                WebDriverWait(driver, 20, 1, ignored_exceptions=None).until(EC.visibility_of_all_elements_located((By.XPATH, '/html/body/div[5]/div/div/div[2]/div/div/dl')))
                time.sleep(1)
            except:
                data_dict_new['location_GRCh37'] = ''
                data_dict_new['location_GRCh38'] = ''
                data_dict_new['Location'] = re.sub(r'\n', ' ', data_dict_new['Location'])
                data_dict_new['Annotations'] = re.sub(r':\n', ': ', data_dict_new['Annotations'])
                result_dicts.append(data_dict_new)
            else:
                # 定位到弹出框
                data_dict_new['location_GRCh37'] = driver.find_element(by=By.XPATH, value='/html/body/div[5]/div/div/div[2]/div/div/dl/dt[1]').text + driver.find_element(by=By.XPATH, value='/html/body/div[5]/div/div/div[2]/div/div/dl/dd[1]').text
                data_dict_new['location_GRCh38'] = driver.find_element(by=By.XPATH, value='/html/body/div[5]/div/div/div[2]/div/div/dl/dt[2]').text + driver.find_element(by=By.XPATH, value='/html/body/div[5]/div/div/div[2]/div/div/dl/dd[2]').text
                data_dict_new['Location'] = re.sub(r'\n', ' ', data_dict_new['Location'])
                data_dict_new['Annotations'] = re.sub(r':\n', ': ', data_dict_new['Annotations'])
                result_dicts.append(data_dict_new)
                # 关闭弹窗
                driver.find_element(by=By.XPATH, value='/html/body/div[5]/div/div/div[1]/button').click()
        # 关闭当前标签页（第二页)
        driver.close()
        driver.switch_to.window(all_handles[0])
        #time.sleep(random.randint(1,3))
        #break
        #time.sleep(10)
        # 关闭子窗口回到原窗口
    return driver, result_dicts, get_list

def main(input, output):
    # driver=webdriver.Chrome(r"C:\Program Files\Google\Chrome\Application\chromedriver.exe")
    driver = webdriver.Chrome(options=options)
    #driver=webdriver.Edge()
    driver.maximize_window()
    #OF = open(output, 'w', encoding='utf-8')
    #newline = '\t'.join(["DECIPHER Patient", "Sex", "chromosome_hg19", "startCoord_hg19", "endCoord_hg19", "chromosome", "startCoord", "endCoord", "Ref", "Alt", "Type", "Size", "Genes", "Genes Count", "Function", "ENST", "cHGVSnom", "pHGVSnom", "hgvs", "Inheritance", "Genotype", "Pathogenicity", "Contribution", "Phenotype(s)"]) + '\n'
    #OF.write(newline)
    out_headers = []
    start_flag = True
    get_list = []
    if not os.path.exists(f"{output}.xlsx"):
        # 创建一个工作簿
        wb = Workbook()
        sheet = wb.active
    else:
        wb = openpyxl.load_workbook(f"{output}.xlsx")
        sheet = wb['Sheet']
    # 获取工作簿的活动表
    n = sheet.max_row # 多一行？
    #print(n)
    if n >= 2:
        out_headers = [cell.value for cell in sheet[1]]
        # DECIPHER Patient __ Location __ Type
        for row in sheet[2:n]:
            data_dict = dict(zip(out_headers, [cell.value for cell in row]))
            get_list.append('__'.join([data_dict[key] for key in ['DECIPHER Patient', 'Location_all']]))
    with open(input, 'r', encoding='utf-8') as IF:
        for line in IF:
            result_dicts = []
            if line.strip().startswith('#') or line.strip().startswith('Assembly') or not line.strip(): continue
            assembly, chrom, start, end = line.strip().split('\t')
            chrom = chrom.strip('chr')
            search_key = assembly + ':' + chrom + ':' + start + '-' + end
            logger.info(f"Search DECIPHER {search_key}")
            driver.get("https://www.deciphergenomics.org/")
            WebDriverWait(driver, 5, 0.5, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value="//*[contains(@id, 'searchFormInput')]"))
            search_element = driver.find_element(by=By.XPATH, value='//*[@id="searchFormInput"]')
            search_element.clear()
            search_element.send_keys(search_key)
            entry_element =  driver.find_element(by=By.XPATH, value='//*[@id="__layout"]/div/div/div[1]/div/div/div[2]/div/form/div[1]/div[1]/div[2]/div/span[2]/button')
            driver.execute_script("arguments[0].click();", entry_element)
            time.sleep(random.randint(3, 10))
            # 提示窗口
            try:
                WebDriverWait(driver, 50, 3, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value='/html/body/div[3]/div/div/div[3]/button'))
                #driver.find_element(by=By.XPATH, value='/html/body/div[3]/div/div/div[3]/button'):
                driver.find_element(by=By.XPATH, value='/html/body/div[3]/div/div/div[3]/button').click()
            except:
                pass
            try:
                # 页面条数选择100 不固定有 不要了
                #time.sleep(100)
                # WebDriverWait(driver, 10, 1, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div/div[1]/div/label/select'))
                # //*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div/div[1]/div/label/select
                # Select(driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div/div[1]/div/label/select')).select_by_visible_text('100')
                WebDriverWait(driver, 20, 1, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/h4'))
                # //*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/h4
                variant_header = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/h4').text
                # Variants: 1 to 10 of 4145
                num_start, num_end, total_num = re.findall(r'(\d+) to (\d+) of (\d+)', variant_header)[0][:]
                logger.info(f"Search DECIPHER {search_key}, get {total_num} records.")
                # 第一页获取内容
                driver, result_dicts, get_list = dealing(driver, search_key, num_start, num_end, total_num, get_list)
                #result_dicts += result_dict
                for result_dict in result_dicts:
                    if n == 1:
                        out_headers = [i for i in result_dict.keys()]
                        sheet.append(out_headers)
                    out_values = [result_dict[key] for key in out_headers]
                    sheet.append(out_values)
                    n += 1
                wb.save(f"{output}.xlsx")
                # print(result_dicts[0][0])
                # return
                while True:
                    try:
                        # 重新打开文件
                        wb = openpyxl.load_workbook(f"{output}.xlsx")
                        # 获取工作簿的活动表
                        sheet = wb['Sheet']
                        # 等待页面
                        #driver.switch_to.window(driver.window_handles[-1])
                        #driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')
                        page_element = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]')
                        next_element = page_element.find_element(by=By.XPATH, value= './/*[contains(text(), "Next")]')
                        action_chains = ActionChains(driver)
                        action_chains.move_to_element(next_element).click().perform()
                        #print(driver.window_handles)
                        WebDriverWait(driver, 10, 1, ignored_exceptions=None).until(lambda x: x.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/h4'))
                    except Exception as e:
                        print(e)
                        traceback.print_exc()
                        break
                    else:
                        variant_header = driver.find_element(by=By.XPATH, value='//*[@id="content"]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/h4').text
                        # Variants: 1 to 10 of 4145
                        num_start, num_end, total_num = re.findall(r'(\d+) to (\d+) of (\d+)', variant_header)[0][:]
                        driver, result_dicts, get_list = dealing(driver, search_key, num_start, num_end, total_num, get_list)
                        # result_dicts += result_dict
                        for result_dict in result_dicts:
                            out_values = [result_dict[key] for key in out_headers]
                            sheet.append(out_values)
                            n+=1
                        wb.save(f"{output}.xlsx")
                        break
                        if num_end == total_num:
                            break
            except Exception as e:
                print(e)
                traceback.print_exc()
                logger.error(f"Search DECIPHER {search_key}, error")
                break # 失败终止
                #continue
            #time.sleep(100)
    logger.info(f"get total DECIPHER records: {n-1}")
    """
    n = 0
    out_headers = []
    if result_dicts:
        for result_dict in result_dicts:
            if n == 0: 
                out_headers = result_dict.keys()
                OF.write('\t'.join(out_headers)+'\n')
            out_values = [re.sub(r'\n', '__line__', result_dict[key]) for key in out_headers]
            OF.write('\t'.join(out_values)+'\n')
            n += 1
    logger.info(f"get total DECIPHER records: {n}")
 
    # 创建一个工作簿
    wb = Workbook()
    # 获取工作簿的活动表
    sheet = wb.active
    sheet.append([i for i in out_headers])
    for result_dict in result_dicts:
        out_values = [result_dict[key] for key in out_headers]
        sheet.append(out_values)
    wb.save(f"{output}.xlsx")
    """
    #OF.close()
    driver.close()

if __name__ == '__main__':
    #input = 'input.txt'
    #output = 'output.tsv'
    main(*sys.argv[1:])
